#!/usr/bin/env python3
"""
Build upload_queue.json from a finalized preview spreadsheet.

Usage:
  python3 build_upload_queue.py <xlsx_path> <video_dir> <connection_id> <out_json> \\
      [--privacy public|unlisted|private] [--order sequential|shuffle] \\
      [--batch-size N] [--seed N]

Reads every row of the spreadsheet (Filename, Title, Description), resolves
each file's actual size on disk, and writes a queue file the scheduled
upload task consumes N items at a time.

Rows with a filename that no longer exists on disk are skipped with a warning --
this keeps a stale spreadsheet from producing a queue that points at nothing.
"""
import sys
import os
import json
import random
from openpyxl import load_workbook


def get_arg(flag, default=None):
    """Get argument value by flag name."""
    if flag in sys.argv:
        idx = sys.argv.index(flag)
        if idx + 1 < len(sys.argv):
            return sys.argv[idx + 1]
    return default


def main():
    if len(sys.argv) < 5:
        print("Usage: python3 build_upload_queue.py <xlsx_path> <video_dir> <connection_id> <out_json> \\")
        print("    [--privacy public|unlisted|private] [--order sequential|shuffle] \\")
        print("    [--batch-size N] [--seed N]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    video_dir = sys.argv[2]
    connection_id = sys.argv[3]
    out_path = sys.argv[4]

    privacy = get_arg("--privacy", "public")
    order = get_arg("--order", "shuffle")
    batch_size = int(get_arg("--batch-size", "3"))
    seed = get_arg("--seed")

    # Validate inputs
    if privacy not in ("public", "unlisted", "private"):
        print(f"Error: Invalid privacy status: {privacy}", file=sys.stderr)
        sys.exit(1)

    if order not in ("sequential", "shuffle"):
        print(f"Error: Invalid order: {order}", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(xlsx_path):
        print(f"Error: Spreadsheet not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    if not os.path.isdir(video_dir):
        print(f"Error: Video directory not found: {video_dir}", file=sys.stderr)
        sys.exit(1)

    # Load spreadsheet
    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        print(f"Error: Could not read spreadsheet: {e}", file=sys.stderr)
        sys.exit(1)

    ws = wb.active

    # Extract items from spreadsheet
    items = []
    missing = []
    empty_metadata = []

    for r in range(2, ws.max_row + 1):
        fn = ws.cell(row=r, column=3).value
        if not fn:
            continue

        path = os.path.join(video_dir, fn)
        if not os.path.exists(path):
            missing.append(fn)
            continue

        title = (ws.cell(row=r, column=5).value or "").strip()
        desc = (ws.cell(row=r, column=6).value or "").strip()

        # Check for missing required metadata
        if not title:
            empty_metadata.append({"filename": fn, "missing": "title"})
        if not desc:
            empty_metadata.append({"filename": fn, "missing": "description"})

        items.append({
            "filename": fn,
            "path": path,
            "title": title,
            "description": desc,
            "fileSizeBytes": os.path.getsize(path),
            "uploaded": False,
            "attempts": 0,
            "videoId": None,
            "error": None,
        })

    # Apply ordering
    if order == "shuffle":
        if seed is not None:
            random.seed(int(seed))
        random.shuffle(items)

    # Build queue
    queue = {
        "connectionId": connection_id,
        "privacyStatus": privacy,
        "batchSize": batch_size,
        "items": items,
    }

    # Save queue
    with open(out_path, "w") as f:
        json.dump(queue, f, ensure_ascii=False, indent=1)

    # Report
    result = {
        "output": out_path,
        "queued": len(items),
        "missing_files_skipped": missing,
        "empty_title_or_desc": [
            i["filename"] for i in items if not i["title"] or not i["description"]
        ],
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))

    # Warn about empty metadata
    if empty_metadata:
        print("\n⚠️  WARNING: Some videos have missing metadata:", file=sys.stderr)
        for item in empty_metadata:
            print(f"   {item['filename']}: missing {item['missing']}", file=sys.stderr)
        print("\nThese videos will be uploaded, but should be filled in before proceeding.", file=sys.stderr)


if __name__ == "__main__":
    main()
