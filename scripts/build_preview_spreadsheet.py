#!/usr/bin/env python3
"""
Build (or refresh) the thumbnail preview spreadsheet for a folder of .mp4 reels.

Usage:
  python3 build_preview_spreadsheet.py <video_dir> <output_xlsx> [--thumb-cache DIR]

Behavior:
- Scans <video_dir> for *.mp4 files.
- Generates a thumbnail for each via macOS Quick Look (qlmanage), cached so
  reruns are fast and don't regenerate thumbnails for files already done.
- If <output_xlsx> already exists, preserves any Title/Description text the
  user already typed in, keyed by filename, and drops rows whose file no
  longer exists. This is a REBUILD, not an in-place edit -- openpyxl does not
  move embedded images when you delete/insert rows, so the only safe way to
  add/remove rows without corrupting the image layout is to rebuild the sheet
  fresh from a merged data set each time.
"""
import sys
import os
import re
import json
import subprocess
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

ROW_H = 90
THUMB_TARGET_W = 160


def parse_date(fn):
    """Extract date from filename if present (YYYY-MM-DD format)."""
    m = re.search(r'(\d{4}-\d{2}-\d{2})', fn)
    return m.group(1) if m else ''


def ensure_thumbnail(video_path, cache_dir):
    """Generate or retrieve cached thumbnail for a video file."""
    fn = os.path.basename(video_path)
    small_path = os.path.join(cache_dir, fn + ".small.png")
    if os.path.exists(small_path):
        return small_path

    raw_path = os.path.join(cache_dir, fn + ".png")
    if not os.path.exists(raw_path):
        # Use macOS Quick Look to generate thumbnail
        subprocess.run(
            ["qlmanage", "-t", "-s", "400", "-o", cache_dir, video_path],
            capture_output=True,
            check=False,
        )

    if not os.path.exists(raw_path):
        return None

    # Resize to target width
    im = PILImage.open(raw_path).convert("RGB")
    w, h = im.size
    new_h = int(h * THUMB_TARGET_W / w)
    im = im.resize((THUMB_TARGET_W, new_h))
    im.save(small_path)
    return small_path


def load_existing_text(xlsx_path):
    """Return {filename: (title, description)} from a prior version of the sheet, if any."""
    if not os.path.exists(xlsx_path):
        return {}
    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        print(f"Warning: Could not read existing spreadsheet: {e}", file=sys.stderr)
        return {}

    ws = wb.active
    out = {}
    for r in range(2, ws.max_row + 1):
        fn = ws.cell(row=r, column=3).value
        title = ws.cell(row=r, column=5).value
        desc = ws.cell(row=r, column=6).value
        if fn:
            out[fn] = (title or "", desc or "")
    return out


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 build_preview_spreadsheet.py <video_dir> <output_xlsx> [--thumb-cache DIR]")
        sys.exit(1)

    video_dir = sys.argv[1]
    out_path = sys.argv[2]
    cache_dir = None

    if "--thumb-cache" in sys.argv:
        cache_dir = sys.argv[sys.argv.index("--thumb-cache") + 1]
    else:
        cache_dir = os.path.join(video_dir, ".preview_thumbs_cache")

    # Validate video directory
    if not os.path.isdir(video_dir):
        print(f"Error: Video directory not found: {video_dir}", file=sys.stderr)
        sys.exit(1)

    os.makedirs(cache_dir, exist_ok=True)

    # Find all .mp4 files
    files = sorted(f for f in os.listdir(video_dir) if f.lower().endswith(".mp4"))

    if not files:
        print(f"Warning: No .mp4 files found in {video_dir}", file=sys.stderr)

    # Load existing captions from prior run
    existing = load_existing_text(out_path)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reels"

    # Headers
    headers = ["#", "Preview", "Filename", "Date", "Title", "Description"]
    ws.append(headers)

    # Style headers
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="333333")
    for col_idx in range(1, 7):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = {"A": 5, "B": 24, "C": 46, "D": 13, "E": 35, "F": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Borders
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add data rows
    for i, fn in enumerate(files, 1):
        row = i + 1
        ws.row_dimensions[row].height = ROW_H
        title, desc = existing.get(fn, ("", ""))

        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=3, value=fn).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=4, value=parse_date(fn)).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=5, value=title).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=6, value=desc).font = Font(name="Arial", size=10)

        # Apply borders and text wrapping
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            if col in (5, 6):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="center")

        # Add thumbnail
        thumb = ensure_thumbnail(os.path.join(video_dir, fn), cache_dir)
        if thumb:
            pil = PILImage.open(thumb)
            w, h = pil.size
            img = XLImage(thumb)
            target_h_px = int(ROW_H * 1.333)
            scale = target_h_px / h
            img.width = int(w * scale)
            img.height = target_h_px
            ws.add_image(img, f"B{row}")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save
    wb.save(out_path)

    # Report
    dropped = sorted(set(existing) - set(files))
    result = {
        "output": out_path,
        "video_count": len(files),
        "preserved_captions": sum(1 for f in files if existing.get(f, ("", ""))[0] or existing.get(f, ("", ""))[1]),
        "dropped_from_previous_sheet": dropped,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
